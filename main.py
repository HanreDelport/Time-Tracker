import sys
from PyQt6.QtWidgets import (QApplication, QMainWindow, QDialog, QMessageBox, QFileDialog,
                             QTreeWidgetItem, QPushButton, QHBoxLayout, QWidget, QInputDialog, QMenu)
from PyQt6 import uic
from database_manager import DatabaseManager
from PyQt6.QtCore import QTimer, Qt
from datetime import datetime
from PyQt6.QtGui import QCloseEvent, QIcon, QBrush, QColor
import csv
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import ctypes


import traceback
import logging

# Setup logging to file
logging.basicConfig(
    filename='time_tracker_debug.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Also log uncaught exceptions
def exception_hook(exctype, value, tb):
    logging.error('Uncaught exception:', exc_info=(exctype, value, tb))
    sys.__excepthook__(exctype, value, tb)

sys.excepthook = exception_hook

# ===== GET RESOURCE PATH =====

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and PyInstaller """
    try:
        base_path = sys._MEIPASS  # PyInstaller temp folder
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

class TimeTrackerApp(QMainWindow):

    #INIT FUNCTION

    def __init__(self):
        super().__init__()
        
        # Set window title
        self.setWindowTitle("Time Tracker")

        # Set the window icon
        self.setWindowIcon(QIcon("assets/stopwatch.png")) 

        # Load the UI file
        uic.loadUi(resource_path('ui/main_window.ui'), self)
        
         # Set window icon
        icon_path = resource_path('assets/stopwatch.ico')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.projectTreeWidget.setUniformRowHeights(True)
        self.projectTreeWidget.setIndentation(18)
        
        # Initialize database manager
        self.db = DatabaseManager()

        # Timer for updating running task
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_running_task)
        self.timer.start(1000)  # Update every 1 second
        
        # Track the currently running task
        self.running_task_id = None
        self.running_task_item = None
        self.task_start_time = None
        self.task_elapsed_before_start = 0  # Seconds already accumulated
        
        # Connect toolbar actions to methods
        self.actionAddProject.triggered.connect(self.add_project)
        self.actionExportCSV.triggered.connect(self.export_to_csv)
        self.actionActionExportExcel.triggered.connect(self.export_to_excel)
        
        # Load projects into the tree
        self.load_projects()
        self.setup_tree_context_menu()

        # Set column widths
        self.projectTreeWidget.setColumnWidth(0, 300)  # Name column
        self.projectTreeWidget.setColumnWidth(1, 100)  # Time column
        self.projectTreeWidget.setColumnWidth(2, 250)  # Actions column
        self.projectTreeWidget.setColumnWidth(3, 100)  # Status column
        
        print("App initialized successfully!")

    # ===== PROJECT METHODS =====

    def add_project(self):
        """Handler for Add Project button"""
        try:
            logging.info("add_project called")
            
            running_task = self.db.get_running_task()
            if running_task:
                QMessageBox.warning(
                    self, 
                    "Task Running", 
                    f"Please pause or finish the currently running task first:\n{running_task[2]}"
                )
                return

            logging.info("About to load dialog UI")
            
            # Load the dialog UI
            dialog = QDialog(self)
            uic.loadUi(resource_path('ui/add_project_dialog.ui'), dialog)
            
            logging.info("Dialog UI loaded successfully")
            
            dialog.buttonBox.accepted.connect(dialog.accept)
            dialog.buttonBox.rejected.connect(dialog.reject)
            
            logging.info("About to show dialog")
            
            # Show the dialog and wait for user response
            if dialog.exec() == QDialog.DialogCode.Accepted:
                # Get the project name from the line edit
                project_name = dialog.projectNameLineEdit.text().strip()
                
                logging.info(f"Dialog accepted with name: {project_name}")
                
                if project_name:
                    # Add to database
                    project_id = self.db.add_project(project_name)
                    print(f"Added project: {project_name} (ID: {project_id})")

                    QMessageBox.information(
                        dialog,
                        "Project Added",
                        f"The project '{project_name}' has been successfully added!"
                    )
                    
                    # Reload the tree
                    self.load_projects()
                else:
                    QMessageBox.warning(self, "Error", "Project name cannot be empty!")
            else:
                reply = QMessageBox.question(
                    dialog,
                    "Cancel",
                    "Are you sure you want to cancel adding the project?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

                if reply == QMessageBox.StandardButton.Yes:
                    dialog.reject()
                    
        except Exception as e:
            logging.error(f"Error in add_project: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")


    def rename_project(self, project_id, old_name):
         # Check if another task is already running
        running_task = self.db.get_running_task()
        if running_task:
            QMessageBox.warning(
                self, 
                "Task Already Running", 
                f"Please pause or finish the currently running task first:\n{running_task[2]}"
            )
            return

        """Rename a project"""
        new_name, ok = QInputDialog.getText(
            self,
            "Rename Project",
            "Enter new project name:",
            text=old_name
        )
        
        if ok and new_name.strip():
            self.db.rename_project(project_id, new_name.strip())
            self.load_projects()
            print(f"Renamed project {project_id} to '{new_name}'")

    def delete_project(self, project_id, project_name):
        """Delete a project"""
        # Check if another task is already running
        running_task = self.db.get_running_task()
        if running_task:
            QMessageBox.warning(
                self, 
                "Task Running", 
                f"Please pause or finish the currently running task first:\n{running_task[2]}"
            )
            return
        
        reply = QMessageBox.question(
            self,
            "Delete Project",
            f"Are you sure you want to delete the project '{project_name}'?\n\n"
            "This will also delete all tasks in this project.\n"
            "This action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_project(project_id)
            self.load_projects()
            print(f"Deleted project {project_id}")

    # ===== TASK METHODS =====      
   
    def add_task_to_project(self, project_id, project_name):
        """Handler for adding a task to a specific project"""
        # Check if another task is already running
        running_task = self.db.get_running_task()
        if running_task:
            QMessageBox.warning(
                self, 
                "Task Running", 
                f"Please pause or finish the currently running task first:\n{running_task[2]}"
            )
            return
        
        # Load the dialog UI
        dialog = QDialog(self)
        uic.loadUi(resource_path('ui/add_task_dialog.ui'), dialog)
        dialog.buttonBox.accepted.connect(dialog.accept)
        dialog.buttonBox.rejected.connect(dialog.reject)
        
        # Update dialog title to show which project
        dialog.setWindowTitle(f"Add Task to {project_name}")
        
        # Show the dialog and wait for user response
        result = dialog.exec()
        print("Dialog result:", result)

        if result == QDialog.DialogCode.Accepted:
            # Get the task name from the line edit
            task_name = dialog.taskNameLineEdit.text().strip()
            
            if task_name:
                # Add to database
                task_id = self.db.add_task(project_id, task_name)
                print(f"Added task: {task_name} to project ID {project_id} (Task ID: {task_id})")
                
                # Reload the tree
                self.load_projects()
            else:
                QMessageBox.warning(self, "Error", "Task name cannot be empty!")
        else:            
            reply = QMessageBox.question(
                dialog,
                "Cancel",
                "Are you sure you want to cancel adding the project?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                dialog.reject()  # Closes the dialog without saving anything

    def create_task_buttons(self, task_item, task_id, is_finished, is_running):
        """Create action buttons for a task"""

        # Create a widget to hold the buttons
        button_widget = QWidget()
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(2, 2, 2, 2)
        button_widget.setLayout(button_layout)
        
        if is_finished:
            # Show only Reopen button for finished tasks
            reopen_btn = QPushButton("Reopen")            
            reopen_btn.setObjectName("primary")
            reopen_btn.clicked.connect(lambda: self.reopen_task(task_id))
            button_layout.addWidget(reopen_btn)
        else:
            # Show Pause and Finish buttons for active tasks
            if is_running:
                pause_btn = QPushButton("Pause")           
                pause_btn.setObjectName("primary")
                pause_btn.clicked.connect(lambda: self.pause_task(task_id))
                button_layout.addWidget(pause_btn)

                finish_btn = QPushButton("Finish")           
                finish_btn.setObjectName("primary")
                finish_btn.clicked.connect(lambda: self.finish_task(task_id))
                button_layout.addWidget(finish_btn)
            else:
                start_btn = QPushButton("Start")                
                start_btn.setObjectName("primary")
                start_btn.clicked.connect(lambda: self.start_task(task_id))
                button_layout.addWidget(start_btn)
            
            
        
        return button_widget

    def start_task(self, task_id):
        """Start a task timer"""
        # Check if another task is already running
        running_task = self.db.get_running_task()
        if running_task:
            QMessageBox.warning(
                self, 
                "Task Already Running", 
                f"Please pause or finish the currently running task first:\n{running_task[2]}"
            )
            return
        
        # Get current task's total seconds from database
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT total_seconds FROM tasks WHERE id = ?', (task_id,))
        result = cursor.fetchone()
        conn.close()
        
        # Start the task
        self.running_task_id = task_id
        self.task_start_time = datetime.now()
        self.task_elapsed_before_start = result[0] if result else 0
        
        # Mark task as running in database
        self.db.start_task(task_id)
        
         # Refresh UI
        self.load_projects()
        
        # Find and store reference to the running task item
        self.find_and_store_running_task_item(task_id)
        
        print(f"Started task {task_id}")

    def pause_task(self, task_id):
        """Pause a task timer"""
        if self.running_task_id != task_id:
            return
        
        # Calculate final time
        current_time = datetime.now()
        elapsed_seconds = int((current_time - self.task_start_time).total_seconds())
        total_seconds = self.task_elapsed_before_start + elapsed_seconds
        
        # Update database
        self.db.update_task_time(task_id, total_seconds)
        self.db.pause_task(task_id)
        
        # Stop tracking
        self.running_task_id = None
        self.task_start_time = None
        self.task_elapsed_before_start = 0
        self.running_task_item = None
        
        # Refresh UI
        self.load_projects()
        print(f"Paused task {task_id}")

    def finish_task(self, task_id):
        """Finish a task"""
        # If task is running, pause it first to save the time
        if self.running_task_id == task_id:
            # Calculate final time
            current_time = datetime.now()
            elapsed_seconds = int((current_time - self.task_start_time).total_seconds())
            total_seconds = self.task_elapsed_before_start + elapsed_seconds
            
            # Update database with final time
            self.db.update_task_time(task_id, total_seconds)
            
            # Stop tracking
            self.running_task_id = None
            self.task_start_time = None
            self.task_elapsed_before_start = 0
            self.running_task_item = None
        
        
        # Mark as finished AND not running
        self.db.finish_task(task_id)
        self.db.pause_task(task_id)  # Make sure it's not marked as running
        
        # Refresh UI
        self.load_projects()
        print(f"Finished task {task_id}")

    def reopen_task(self, task_id):
        """Reopen a finished task"""
        running_task = self.db.get_running_task()
        if running_task:
            QMessageBox.warning(
                self, 
                "Task Running", 
                f"Please pause or finish the currently running task first:\n{running_task[2]}"
            )
            return

        self.db.reopen_task(task_id)
        self.db.pause_task(task_id)  # Make sure it starts as paused, not running
        self.load_projects()
        print(f"Reopened task {task_id}")

    def find_and_store_running_task_item(self, task_id):
        """Find the tree widget item for the running task"""
        # Search through all projects
        for i in range(self.projectTreeWidget.topLevelItemCount()):
            project_item = self.projectTreeWidget.topLevelItem(i)
            
            # Search through all tasks in this project
            for j in range(project_item.childCount()):
                task_item = project_item.child(j)
                
                # Check if this is our task
                if task_item.data(0, 1) == task_id:
                    self.running_task_item = task_item
                    return

    def rename_task(self, task_id, old_name):
        """Rename a task"""
        running_task = self.db.get_running_task()
        if running_task:
            QMessageBox.warning(
                self, 
                "Task Running", 
                f"Please pause or finish the currently running task first:\n{running_task[2]}"
            )
            return

        new_name, ok = QInputDialog.getText(
            self,
            "Rename Task",
            "Enter new task name:",
            text=old_name
        )
        
        if ok and new_name.strip():
            self.db.rename_task(task_id, new_name.strip())
            self.load_projects()
            print(f"Renamed task {task_id} to '{new_name}'")

    def delete_task(self, task_id, task_name):
        """Delete a task"""
        running_task = self.db.get_running_task()
        if running_task:
            QMessageBox.warning(
                self, 
                "Task Running", 
                f"Please pause or finish the currently running task first:\n{running_task[2]}"
            )
            return

        # Check if this task is currently running
        if self.running_task_id == task_id:
            QMessageBox.warning(
                self,
                "Cannot Delete",
                "Cannot delete a running task. Please pause or finish it first."
            )
            return
        
        reply = QMessageBox.question(
            self,
            "Delete Task",
            f"Are you sure you want to delete the task '{task_name}'?\n\n"
            "This action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_task(task_id)
            self.load_projects()
            print(f"Deleted task {task_id}")

    # ===== TREE METHODS =====           
    
    def setup_tree_context_menu(self):
        """Setup right-click context menu for the tree"""
        from PyQt6.QtCore import Qt
        self.projectTreeWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.projectTreeWidget.customContextMenuRequested.connect(self.show_context_menu)

    def show_context_menu(self, position):
        """Show context menu when right-clicking on tree items"""
        item = self.projectTreeWidget.itemAt(position)
        if item is None:
            return
        
        menu = QMenu()
        
        # Check if this is a project item (parent is None) or task item (has parent)
        if item.parent() is None:
            # This is a PROJECT item
            project_id = item.data(0, 1)
            project_name = item.text(0)
            
            add_task_action = menu.addAction("Add Task")
            rename_project_action = menu.addAction("Rename Project")
            delete_project_action = menu.addAction("Delete Project")
            
            action = menu.exec(self.projectTreeWidget.viewport().mapToGlobal(position))
            
            if action == add_task_action:
                self.add_task_to_project(project_id, project_name)
            elif action == rename_project_action:
                self.rename_project(project_id, project_name)
            elif action == delete_project_action:
                self.delete_project(project_id, project_name)
        
        else:
            # This is a TASK item
            task_id = item.data(0, 1)
            task_name = item.text(0)
            
            rename_task_action = menu.addAction("Rename Task")
            delete_task_action = menu.addAction("Delete Task")
            
            action = menu.exec(self.projectTreeWidget.viewport().mapToGlobal(position))
            
            if action == rename_task_action:
                self.rename_task(task_id, task_name)
            elif action == delete_task_action:
                self.delete_task(task_id, task_name)

    def get_tree_state(self):
        state = {}
        for i in range(self.projectTreeWidget.topLevelItemCount()):
            item = self.projectTreeWidget.topLevelItem(i)
            project_id = item.data(0, 1)
            state[project_id] = item.isExpanded()
        return state
    
    def restore_tree_state(self, state):
        for i in range(self.projectTreeWidget.topLevelItemCount()):
            item = self.projectTreeWidget.topLevelItem(i)
            project_id = item.data(0, 1)
            if project_id in state:
                item.setExpanded(state[project_id])


        
    def load_projects(self):
        """Load all projects from database into the tree widget"""
        #Get current state of tree
        tree_state = self.get_tree_state()

        # Clear the tree
        self.projectTreeWidget.clear()
        
        # Get all projects from database
        projects = self.db.get_all_projects()
        
        print(f"Loaded {len(projects)} projects from database")
        
        # Add each project to the tree
        for project in projects:
            project_id, project_name = project
            
            # Create a tree item for the project
            project_item = QTreeWidgetItem(self.projectTreeWidget)
            project_item.setText(0, project_name)  # Column 0: Name
            
            # Store the project ID in the item (we'll need this later)
            project_item.setData(0, 1, project_id)  # Store ID in role 1
            
            # Get tasks for this project
            tasks = self.db.get_tasks_for_project(project_id)
            
            # Calculate total time for the project
            total_seconds = sum(task[2] for task in tasks)  # task[2] is total_seconds
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            project_item.setText(1, time_str)  # Column 1: Time
            project_item.setText(3, f"{len(tasks)} task(s)")  # Column 3: Status
            
            # Make project item expandable
            project_item.setExpanded(False)

            # Add tasks under this project
            for task in tasks:
                task_id, task_name, total_seconds, is_finished, is_running = task
                
                # Create a tree item for the task (child of project)
                task_item = QTreeWidgetItem(project_item)
                task_item.setText(0, task_name)  # Column 0: Task name
                
                # Store the task ID in the item
                task_item.setData(0, 1, task_id)
                
                # Format and display time
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                task_item.setText(1, time_str)  # Column 1: Time
                
                # Column 2: Action buttons
                button_widget = self.create_task_buttons(task_item, task_id, is_finished, is_running)
                self.projectTreeWidget.setItemWidget(task_item, 2, button_widget)
                
                # Column 3: Status
                if is_finished:
                    task_item.setText(3, "Finished")
                elif is_running:
                    task_item.setText(3, "Running")
                else:
                    task_item.setText(3, "Paused")

        # Highlight running task
        if self.running_task_id is not None:
            for i in range(self.projectTreeWidget.topLevelItemCount()):
                projectOpen = False
                project_item = self.projectTreeWidget.topLevelItem(i)
                for j in range(project_item.childCount()):
                    task_item = project_item.child(j)
                    task_id = task_item.data(0, 1)
                    
                    if task_id == self.running_task_id:
                        projectOpen = True

                        # Set a light blue background
                        for col in range(self.projectTreeWidget.columnCount()):
                            task_item.setBackground(col, QBrush(QColor("#23cff6")))  # pale blue
                            task_item.setForeground(col, QBrush(QColor("#1e3a8a")))  # dark blue text

                            for col in range(self.projectTreeWidget.columnCount()):
                                font = task_item.font(col)
                                font.setBold(True)
                                task_item.setFont(col, font)

                # Highlight parent project
                if projectOpen:
                    for col in range(self.projectTreeWidget.columnCount()):
                        project_item.setBackground(col, QBrush(QColor("#23cff6"))) 
                        project_item.setForeground(col, QBrush(QColor("#1d4ed8")))

                        for col in range(self.projectTreeWidget.columnCount()):
                                font = project_item.font(col)
                                font.setBold(True)
                                project_item.setFont(col, font)
        
        #Stretch Collumns out with window
        header = self.projectTreeWidget.header()
        for col in range(self.projectTreeWidget.columnCount()):
            header.setSectionResizeMode(col, header.ResizeMode.Stretch)


        self.restore_tree_state(tree_state)


    def update_project_total_time(self, project_item):
        """Update the total time display for a project"""
        total_seconds = 0
        
        # Sum up all child task times
        for i in range(project_item.childCount()):
            task_item = project_item.child(i)
            time_text = task_item.text(1)  # Get time string like "00:05:23"
            
            # Parse the time string
            time_parts = time_text.split(':')
            if len(time_parts) == 3:
                hours, minutes, seconds = map(int, time_parts)
                total_seconds += hours * 3600 + minutes * 60 + seconds
        
        # Update project's time display
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        project_item.setText(1, time_str)

    # ===== TIMER =====

    def update_running_task(self):
        """Called every second to update the running task's time display"""
        if self.running_task_id is None or self.running_task_item is None:
            return
        
        # Calculate elapsed time since start
        current_time = datetime.now()
        elapsed_seconds = int((current_time - self.task_start_time).total_seconds())
        
        # Total time = previous time + current session time
        total_seconds = self.task_elapsed_before_start + elapsed_seconds
        
        # Update only the display (no database write yet)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        
        # Update only the time column in the UI
        self.running_task_item.setText(1, time_str)
        
        # Also update the parent project's total time
        parent_item = self.running_task_item.parent()
        if parent_item:
            self.update_project_total_time(parent_item)

    # ===== HANDLE CLOSING =====

    def closeEvent(self, event: QCloseEvent):
        """Handle window close event"""
        # Check if a task is running
        if self.running_task_id is not None:
            # Show warning
            reply = QMessageBox.warning(
                self,
                "Task Running",
                "A task is currently running. Are you sure you want to close?\n\n"
                "The timer will stop and current progress will be saved.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.No:
                event.ignore()  # Cancel the close
                return
            else:
                # Save the running task before closing
                current_time = datetime.now()
                elapsed_seconds = int((current_time - self.task_start_time).total_seconds())
                total_seconds = self.task_elapsed_before_start + elapsed_seconds
                self.db.update_task_time(self.running_task_id, total_seconds)
                self.db.pause_task(self.running_task_id)
        
        # Accept the close event (actually close the application)
        event.accept()
   
    # ===== EXPORTING =====

    def export_to_csv(self):
        # Check if another task is already running
        running_task = self.db.get_running_task()
        if running_task:
            QMessageBox.warning(
                self, 
                "Task Running", 
                f"Please pause or finish the currently running task first:\n{running_task[2]}"
            )
            return

        """Export all projects and tasks to CSV (.csv)"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to CSV",
            "exports/time_tracker_export.csv",
            "CSV Files (*.csv)"
        )

        if not file_path:
            return

        try:
            with open(file_path, mode='w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                # Header
                headers = ['Project', 'Task', 'Time (HH:MM:SS)', 'Status']
                writer.writerow(headers)

                projects = self.db.get_all_projects()

                for project_id, project_name in projects:
                    tasks = self.db.get_tasks_for_project(project_id)

                    if not tasks:
                        project_time = "00:00:00"
                        project_status = "N/A"
                        writer.writerow([project_name, '', project_time, 'No tasks'])
                    else:
                        # Calculate total project time
                        total_seconds = sum(task[2] for task in tasks)
                        h = total_seconds // 3600
                        m = (total_seconds % 3600) // 60
                        s = total_seconds % 60
                        project_time = f"{h:02d}:{m:02d}:{s:02d}"

                        # Determine project status
                        project_status = "Finished" if all(task[3] for task in tasks) else "Open"

                        # First task in the same row as project
                        task_id, task_name, task_seconds, is_finished, is_running = tasks[0]
                        h = task_seconds // 3600
                        m = (task_seconds % 3600) // 60
                        s = task_seconds % 60
                        time_str = f"{h:02d}:{m:02d}:{s:02d}"

                        if is_finished:
                            status = "Finished"
                        elif is_running:
                            status = "Running"
                        else:
                            status = "Paused"

                        writer.writerow([project_name, task_name, time_str, status])

                        # Remaining tasks
                        for task in tasks[1:]:
                            task_id, task_name, task_seconds, is_finished, is_running = task
                            h = task_seconds // 3600
                            m = (task_seconds % 3600) // 60
                            s = task_seconds % 60
                            time_str = f"{h:02d}:{m:02d}:{s:02d}"

                            if is_finished:
                                status = "Finished"
                            elif is_running:
                                status = "Running"
                            else:
                                status = "Paused"

                            writer.writerow(['', task_name, time_str, status])

                    # Project Total row
                    writer.writerow([
                        "Total:",
                        '',
                        project_time,
                        f"{len(tasks)} task(s), Status: {project_status}"
                    ])

            # Automatically open the CSV file after saving (Windows)
            import os
            os.startfile(file_path)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export CSV:\n{e}")

    def export_to_excel(self):
        """Export all projects and tasks to Excel (.xlsx)"""
        # Check if another task is already running
        running_task = self.db.get_running_task()
        if running_task:
            QMessageBox.warning(
                self, 
                "Task Running", 
                f"Please pause or finish the currently running task first:\n{running_task[2]}"
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            "exports/time_tracker_export.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Time Tracker"

            # Header
            headers = ['Project', 'Task', 'Time (HH:MM:SS)', 'Status']
            ws.append(headers)

            # Bold header
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)

            projects = self.db.get_all_projects()

            for project_id, project_name in projects:
                tasks = self.db.get_tasks_for_project(project_id)

                if not tasks:
                    ws.append([project_name, 'N/A', '00:00:00', 'No tasks'])
                    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
                else:
                    total_seconds = sum(task[2] for task in tasks)
                    h = total_seconds // 3600
                    m = (total_seconds % 3600) // 60
                    s = total_seconds % 60
                    project_time = f"{h:02d}:{m:02d}:{s:02d}"

                    projectFirstLine = True

                    for task_id, task_name, task_seconds, is_finished, is_running in tasks:
                        h = task_seconds // 3600
                        m = (task_seconds % 3600) // 60
                        s = task_seconds % 60
                        time_str = f"{h:02d}:{m:02d}:{s:02d}"

                        # Determine project status
                        if all(task[3] for task in tasks):  
                            project_status = "Finished"
                        else:
                            project_status = "In Progress"  

                        if is_finished:
                            status = "Finished"
                        elif is_running:
                            status = "Running"
                        else:
                            status = "Paused"

                        #Include project name in the first line of the project
                        if (projectFirstLine):
                            ws.append([
                                f"{project_name}",
                                task_name,
                                time_str,
                                status
                            ])
                            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
                            projectFirstLine = False
                        else:
                            ws.append([
                                "",
                                task_name,
                                time_str,
                                status
                            ])


                    # Project summary row
                    ws.append([
                        '',
                        f"{len(tasks)} task(s)",
                        project_time,
                        project_status
                    ])

                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True)
                    
                # Empty row between projects
                ws.append([])

            # Auto-size columns
            for column in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
                ws.column_dimensions[column[0].column_letter].width = max_length + 2

            wb.save(file_path)

            QMessageBox.information(
                self,
                "Export Successful",
                f"Data exported successfully to:\n{file_path}"
            )

            # Open the Excel file
            os.startfile(file_path)

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Failed to export data:\n{str(e)}"
            )


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(resource_path("assets/stopwatch.ico")))
    with open(resource_path("styles/app.qss"), "r") as f:
        app.setStyleSheet(f.read())
    window = TimeTrackerApp()
    window.show()
    sys.exit(app.exec())